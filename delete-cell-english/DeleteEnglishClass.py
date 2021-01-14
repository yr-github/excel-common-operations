from openpyxl import load_workbook
import re


class DeleteEnglishClass():
    def __init__(self, files, sheet_names):
        self.files = files
        self.sheet_names = sheet_names
        return

    def open_excel(self, file):
        wb = load_workbook(filename=file)
        return wb

    def delete_eng(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    unen = re.compile('([\u4e00-\u9fa5].*)')
                    cn = "".join(unen.findall(cell.value))
                    # uncn = re.compile("[^\u4e00-\u9fa5]")
                    # en = "".join(uncn.findall(cell.value))
                    # if cn != '' and en != '':
                    # 修改一下逻辑 实际使用有中文直接复制即可
                    if cn != '':
                        cell.value = cn
                        print(cell)
        return

    def run(self):
        for file in self.files:
            excel = self.open_excel(file)
            for sheet_name in self.sheet_names:
                self.delete_eng(excel[sheet_name])
            copy = file.split('.')[0]+'_copy.xlsx'
            excel.save(copy)
        return
