from openpyxl import load_workbook
from openpyxl.styles import Alignment


class TranslationByHistoryClass():
    def __init__(self, history, files, sheet_names):
        self.history = history
        self.files = files
        self.sheet_names = sheet_names
        return

    def open_excel(self, file):
        wb = load_workbook(filename=file)
        return wb

    def _delete_blank(self, mystr):
        result = mystr.replace(' ', '')
        return result

    # 判断当前单元格是否有历史翻译
    def _is_key(self, key):
        result = None
        key = self._delete_blank(key)
        for history in self.history:
            if key == self._delete_blank(history):
                result = self.history[history]
        return result

    def translate_eng(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    value = self._is_key(cell.value)
                    if isinstance(value, str):
                        cell.value = cell.value + '\n' + value
                        cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                   vertical=cell.alignment.vertical,
                                                   text_rotation=cell.alignment.text_rotation,
                                                   wrap_text=True,
                                                   shrink_to_fit=cell.alignment.shrink_to_fit,
                                                   indent=cell.alignment.indent)
                        print(cell)

        return

    def run(self):
        for file in self.files:
            excel = self.open_excel(file)
            for sheet_name in self.sheet_names:
                self.translate_eng(excel[sheet_name])
            copy = file.split('.')[0]+'_translate.xlsx'
            excel.save(copy)
        return
